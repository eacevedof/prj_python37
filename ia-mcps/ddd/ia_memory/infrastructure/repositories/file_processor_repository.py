import os
from pathlib import Path
from typing import Any, final, Self

from ddd.ia_memory.domain.exceptions import MemoryException


@final
class FileProcessorRepository:
    """Repository for processing binary files into text chunks."""

    _instance: "FileProcessorRepository | None" = None

    def __init__(self) -> None:
        pass

    @classmethod
    def get_instance(cls) -> Self:
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    def process_file(self, file_path: str) -> list[dict[str, Any]]:
        if not os.path.exists(file_path):
            raise MemoryException.file_not_found(file_path)
        ext = Path(file_path).suffix.lower()
        processors = {
            ".pdf": self._process_pdf,
            ".png": self._process_image,
            ".jpg": self._process_image,
            ".jpeg": self._process_image,
            ".mp3": self._process_audio,
            ".wav": self._process_audio,
            ".docx": self._process_docx,
            ".xlsx": self._process_xlsx,
        }
        processor = processors.get(ext)
        if not processor:
            raise MemoryException.unsupported_file_type(ext)
        return processor(file_path)

    def _process_pdf(self, file_path: str) -> list[dict[str, Any]]:
        try:
            import fitz
            doc = fitz.open(file_path)
            chunks = []
            for page_num, page in enumerate(doc):
                text = page.get_text().strip()
                if text:
                    chunks.append({
                        "content": f"[Page {page_num + 1}] {text}",
                        "metadata": {
                            "source_file": file_path,
                            "source_type": "pdf",
                            "page": page_num + 1,
                        }
                    })
            doc.close()
            return chunks
        except ImportError:
            raise MemoryException.file_processing_error(file_path, "PyMuPDF not installed")
        except Exception as e:
            raise MemoryException.file_processing_error(file_path, str(e))

    def _process_image(self, file_path: str) -> list[dict[str, Any]]:
        try:
            from PIL import Image
            from transformers import BlipProcessor, BlipForConditionalGeneration
            processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base")
            model = BlipForConditionalGeneration.from_pretrained("Salesforce/blip-image-captioning-base")
            image = Image.open(file_path)
            inputs = processor(image, return_tensors="pt")
            output = model.generate(**inputs, max_length=200)
            description = processor.decode(output[0], skip_special_tokens=True)
            return [{
                "content": f"[Image] {description}",
                "metadata": {"source_file": file_path, "source_type": "image"}
            }]
        except ImportError:
            raise MemoryException.file_processing_error(file_path, "transformers not installed")
        except Exception as e:
            raise MemoryException.file_processing_error(file_path, str(e))

    def _process_audio(self, file_path: str) -> list[dict[str, Any]]:
        try:
            import whisper
            model = whisper.load_model("base")
            result = model.transcribe(file_path)
            text = result["text"].strip()
            return [{
                "content": f"[Audio] {text}",
                "metadata": {"source_file": file_path, "source_type": "audio"}
            }]
        except ImportError:
            raise MemoryException.file_processing_error(file_path, "whisper not installed")
        except Exception as e:
            raise MemoryException.file_processing_error(file_path, str(e))

    def _process_docx(self, file_path: str) -> list[dict[str, Any]]:
        try:
            from docx import Document
            doc = Document(file_path)
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return [{
                "content": text,
                "metadata": {"source_file": file_path, "source_type": "docx"}
            }]
        except ImportError:
            raise MemoryException.file_processing_error(file_path, "python-docx not installed")
        except Exception as e:
            raise MemoryException.file_processing_error(file_path, str(e))

    def _process_xlsx(self, file_path: str) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            chunks = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(c) if c else "" for c in row)
                    if row_text.replace("|", "").strip():
                        rows.append(row_text)
                if rows:
                    chunks.append({
                        "content": f"[Sheet: {sheet_name}]\n" + "\n".join(rows[:50]),
                        "metadata": {"source_file": file_path, "source_type": "xlsx", "sheet": sheet_name}
                    })
            wb.close()
            return chunks
        except ImportError:
            raise MemoryException.file_processing_error(file_path, "openpyxl not installed")
        except Exception as e:
            raise MemoryException.file_processing_error(file_path, str(e))
